import os
import sqlite3
from datetime import datetime

from openpyxl import load_workbook
from database.repository import upsert_many

BATCH_SIZE = 1000
HISTORY_DB_PATH = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "import_history.db")
)


def _get_history_db_connection():
    return sqlite3.connect(HISTORY_DB_PATH)


def _ensure_import_history_table():
    with _get_history_db_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS import_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_name TEXT,
                imported_records INTEGER,
                import_date TEXT,
                import_time TEXT,
                duration_seconds REAL
            )
            """
        )


def insert_import_history(
    file_name,
    imported_records,
    import_date,
    import_time,
    duration_seconds
):
    _ensure_import_history_table()
    with _get_history_db_connection() as conn:
        conn.execute(
            """
            INSERT INTO import_history (
                file_name,
                imported_records,
                import_date,
                import_time,
                duration_seconds
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (
                file_name,
                imported_records,
                import_date,
                import_time,
                duration_seconds,
            ),
        )


def get_import_history(limit=None):
    _ensure_import_history_table()
    query = (
        "SELECT id, file_name, imported_records, import_date, import_time, duration_seconds "
        "FROM import_history ORDER BY id DESC"
    )
    params = ()

    if limit is not None:
        query += " LIMIT ?"
        params = (limit,)

    with _get_history_db_connection() as conn:
        cursor = conn.execute(query, params)
        return cursor.fetchall()


def _save_import_history(file_path, imported_records, duration_seconds, timestamp):
    insert_import_history(
        file_name=os.path.basename(file_path),
        imported_records=imported_records,
        import_date=timestamp.strftime("%Y-%m-%d"),
        import_time=timestamp.strftime("%H:%M:%S"),
        duration_seconds=duration_seconds,
    )


_ensure_import_history_table()


def import_excel(file_path, progress_callback=None):
    start_time = datetime.now()

    # Workbook Open
    wb = load_workbook(
        filename=file_path,
        read_only=True,
        data_only=True
    )

    ws = wb.active

    # Total Rows
    total_rows = ws.max_row - 1

    # Read Rows
    rows = ws.iter_rows(values_only=True)

    headers = next(rows)

    headers = [
        str(h).strip() if h else ""
        for h in headers
    ]

    column = {
        name: index
        for index, name in enumerate(headers)
    }

    batch = []

    processed = 0

    for row in rows:

        meter_no = str(
            row[column["Meter no."]]
        ).strip()

        if meter_no in ("", "None"):
            continue

        batch.append((
            meter_no,
            row[column["Consumer Number"]],
            row[column["Consumer Name"]],
            row[column["Father name"]],
            row[column["Address"]],
            row[column["Zone"]],
            row[column["Division"]],
            row[column["Subdivision"]],
            row[column["Mobile Number 1"]],
            row[column["Mobile number 2"]],
            row[column["Location"]],
            row[column["Whats app remark"]],
        ))

        processed += 1

        if len(batch) >= BATCH_SIZE:

            upsert_many(batch)

            batch.clear()

            if progress_callback:
                progress_callback(
                    processed,
                    total_rows
                )

    if batch:
        upsert_many(batch)

    wb.close()

    end_time = datetime.now()
    duration_seconds = (end_time - start_time).total_seconds()
    _save_import_history(
        file_path=file_path,
        imported_records=processed,
        duration_seconds=duration_seconds,
        timestamp=end_time,
    )

    if progress_callback:
        progress_callback(
            total_rows,
            total_rows
        )

    return processed